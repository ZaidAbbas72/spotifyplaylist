"""
Excel Export Processor for Spotify Playlist Data
Handles Excel file generation with comprehensive formatting and multiple sheets
"""

import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

logger = logging.getLogger(__name__)

class ExcelProcessor:
    def __init__(self):
        """Initialize Excel processor"""
        self.wb = None
        
    def create_excel(self, playlist_data, tracks_data):
        """Create Excel workbook with playlist and tracks data"""
        try:
            self.wb = Workbook()
            
            # Remove default worksheet
            self.wb.remove(self.wb.active)
            
            # Create multiple sheets
            self._create_summary_sheet(playlist_data, tracks_data)
            self._create_tracks_sheet(tracks_data)
            self._create_audio_features_sheet(tracks_data)
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            self.wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info("Excel file created successfully")
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            raise Exception(f"Failed to create Excel file: {e}")
    
    def _create_summary_sheet(self, playlist_data, tracks_data):
        """Create summary sheet with playlist metadata"""
        ws = self.wb.create_sheet("Playlist Summary", 0)
        
        # Title
        ws['A1'] = "Spotify Playlist Summary"
        ws['A1'].font = Font(size=18, bold=True, color="1DB954")
        ws.merge_cells('A1:D1')
        
        # Playlist metadata
        metadata = [
            ("Playlist Name", playlist_data.get('name', 'Unknown')),
            ("Description", playlist_data.get('description', 'No description')),
            ("Total Saves/Followers", playlist_data.get('followers', 0)),
            ("Number of Songs", playlist_data.get('total_tracks', 0)),
            ("Total Duration", playlist_data.get('total_duration', 'Unknown')),
            ("Spotify URL", playlist_data.get('external_url', '')),
            ("Extracted on", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ]
        
        row = 3
        for label, value in metadata:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Statistics
        ws[f'A{row + 1}'] = "Statistics"
        ws[f'A{row + 1}'].font = Font(size=14, bold=True, color="1DB954")
        row += 2
        
        # Calculate statistics
        total_duration_ms = sum(track.get('duration_ms', 0) for track in tracks_data)
        avg_popularity = sum(track.get('popularity', 0) for track in tracks_data) / len(tracks_data) if tracks_data else 0
        explicit_count = sum(1 for track in tracks_data if track.get('explicit', False))
        
        stats = [
            ("Total Tracks Extracted", len(tracks_data)),
            ("Average Popularity", f"{avg_popularity:.1f}"),
            ("Explicit Tracks", explicit_count),
            ("Tracks with Audio Features", sum(1 for track in tracks_data if track.get('danceability') not in [None, 'N/A'])),
            ("Unique Artists", len(set(track.get('artists', 'Unknown') for track in tracks_data)))
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Style the sheet
        self._style_worksheet(ws, "Summary")
    
    def _create_tracks_sheet(self, tracks_data):
        """Create detailed tracks sheet"""
        ws = self.wb.create_sheet("Track Details")
        
        # Headers
        headers = [
            'Track #', 'Track Name', 'Artist(s)', 'Album Name', 'Duration (mm:ss)',
            'Date Added', 'Release Year', 'Popularity', 'Explicit', 'Streams'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1DB954", end_color="1DB954", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row, track in enumerate(tracks_data, 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=track.get('track_name', 'Unknown'))
            ws.cell(row=row, column=3, value=track.get('artists', 'Unknown Artist'))
            ws.cell(row=row, column=4, value=track.get('album_name', 'Unknown Album'))
            ws.cell(row=row, column=5, value=track.get('duration', '0:00'))
            ws.cell(row=row, column=6, value=track.get('date_added', 'Unknown'))
            ws.cell(row=row, column=7, value=track.get('release_year', 'Unknown'))
            ws.cell(row=row, column=8, value=track.get('popularity', 0))
            ws.cell(row=row, column=9, value='Yes' if track.get('explicit', False) else 'No')
            ws.cell(row=row, column=10, value=track.get('streams', 'N/A'))
        
        self._style_worksheet(ws, "Tracks")
    
    def _create_audio_features_sheet(self, tracks_data):
        """Create audio features sheet"""
        ws = self.wb.create_sheet("Audio Features")
        
        # Check if we have audio features
        has_features = any(track.get('danceability') not in [None, 'N/A'] for track in tracks_data)
        
        if not has_features:
            ws['A1'] = "Audio Features Not Available"
            ws['A1'].font = Font(size=14, bold=True, color="FF0000")
            ws['A3'] = "Audio features require extended API permissions or may not be available for all tracks."
            return
        
        # Headers
        headers = [
            'Track #', 'Track Name', 'Danceability', 'Energy', 'Valence', 'Tempo',
            'Acousticness', 'Instrumentalness', 'Liveness', 'Speechiness',
            'Loudness', 'Key', 'Mode', 'Time Signature'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1DB954", end_color="1DB954", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row, track in enumerate(tracks_data, 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=track.get('track_name', 'Unknown'))
            ws.cell(row=row, column=3, value=self._format_feature(track.get('danceability')))
            ws.cell(row=row, column=4, value=self._format_feature(track.get('energy')))
            ws.cell(row=row, column=5, value=self._format_feature(track.get('valence')))
            ws.cell(row=row, column=6, value=self._format_feature(track.get('tempo')))
            ws.cell(row=row, column=7, value=self._format_feature(track.get('acousticness')))
            ws.cell(row=row, column=8, value=self._format_feature(track.get('instrumentalness')))
            ws.cell(row=row, column=9, value=self._format_feature(track.get('liveness')))
            ws.cell(row=row, column=10, value=self._format_feature(track.get('speechiness')))
            ws.cell(row=row, column=11, value=self._format_feature(track.get('loudness')))
            ws.cell(row=row, column=12, value=self._format_feature(track.get('key')))
            ws.cell(row=row, column=13, value=self._format_feature(track.get('mode')))
            ws.cell(row=row, column=14, value=self._format_feature(track.get('time_signature')))
        
        self._style_worksheet(ws, "Audio Features")
    
    def _format_feature(self, value):
        """Format audio feature values"""
        if value is None or value == 'N/A':
            return 'N/A'
        if isinstance(value, (int, float)):
            return round(value, 3)
        return str(value)
    
    def _style_worksheet(self, ws, sheet_type):
        """Apply styling to worksheet"""
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        if sheet_type != "Summary":
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.border = thin_border
        
        # Freeze header row for data sheets
        if sheet_type in ["Tracks", "Audio Features"]:
            ws.freeze_panes = "A2"
        
        # Alternate row colors for better readability
        if sheet_type in ["Tracks", "Audio Features"]:
            light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            for row_num in range(3, ws.max_row + 1, 2):
                for col_num in range(1, ws.max_column + 1):
                    ws.cell(row=row_num, column=col_num).fill = light_fill